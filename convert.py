"""导入题库到 SQLite（支持 .xls / .xlsx）"""
import json
import os
import re
import sqlite3

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'data', 'shuati.db')

LABEL_MAP = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F'}


def parse_legacy(row):
    """旧版 xls 格式：category | type | answer | content(含选项)"""
    category = str(row[0]).strip()
    qtype = str(row[1]).strip()
    answer = str(row[2]).strip().lower()
    content = str(row[3]).strip()

    options, answer_indices, answer_text = [], [], ""
    pure_content = content

    if qtype == "判断题":
        options = ["正确", "错误"]
        answer_indices = [0] if answer == 'y' else [1]
        answer_text = "正确" if answer == 'y' else "错误"
    elif qtype in ("单选题", "多选题"):
        lines = content.split('\n')
        pure_content = lines[0].strip()
        opts, labels = [], []
        for line in lines[1:]:
            m = re.match(r'^([A-Da-d])[、．.．\s]\s*(.+)', line.strip())
            if m:
                labels.append(m.group(1).upper())
                opts.append(m.group(2).strip())
        options = opts
        label_lower = [l.lower() for l in labels]
        if qtype == "单选题":
            if answer in label_lower:
                idx = label_lower.index(answer)
                answer_indices = [idx]
                answer_text = f"{labels[idx]}、{opts[idx]}"
        else:
            indices = sorted([label_lower.index(ch) for ch in answer if ch in label_lower])
            answer_indices = indices
            answer_text = "、".join([f"{labels[i]}、{opts[i]}" for i in indices])

    return category, qtype, pure_content, options, answer_indices, answer_text, answer


def parse_xlsx(row):
    """新版 xlsx 格式：type | number | content | optA | optB | optC | optD | answer | category | ..."""
    qtype = str(row[0]).strip() if row[0] else ""
    content = str(row[2]).strip() if row[2] else ""
    opts = [str(row[3]).strip() if row[3] is not None else "",
            str(row[4]).strip() if row[4] is not None else "",
            str(row[5]).strip() if row[5] is not None else "",
            str(row[6]).strip() if row[6] is not None else ""]
    raw_answer = str(row[7]).strip().upper() if row[7] else ""
    category = str(row[8]).strip() if row[8] else ""

    options, answer_indices, answer_text = [], [], ""

    if qtype == "判断题":
        options = ["正确", "错误"]
        answer_indices = [0] if raw_answer in ("Y", "正确", "对") else [1]
        answer_text = options[answer_indices[0]]
    elif qtype in ("单选题", "多选题"):
        options = [o for o in opts if o]
        for ch in raw_answer:
            for i, label in LABEL_MAP.items():
                if i < len(options) and ch == label:
                    answer_indices.append(i)
                    break
        answer_indices = sorted(answer_indices)
        answer_text = "、".join([f"{LABEL_MAP[i]}、{options[i]}" for i in answer_indices])

    return category, qtype, content, options, answer_indices, answer_text, raw_answer


def import_bank(bank_name, filepath):
    """导入一个题库文件"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(filepath)
        sheet = wb.sheet_by_name('Sheet1')
        rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(1, sheet.nrows)]
        parser = parse_legacy
    elif ext == '.xlsx':
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(3, ws.max_row + 1)]
        parser = parse_xlsx
    else:
        print(f"  跳过不支持格式: {ext}")
        return 0

    conn = sqlite3.connect(DB_PATH)

    total = 0
    cats, types = {}, {}

    for row in rows:
        try:
            category, qtype, content, options, answer_indices, answer_text, raw_answer = parser(row)
        except Exception:
            continue
        if not qtype or not content:
            continue

        conn.execute('''
            INSERT INTO questions (bank, category, type, content, options, answer, answer_text, raw_answer)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            bank_name, category, qtype, content,
            json.dumps(options, ensure_ascii=False),
            json.dumps(answer_indices),
            answer_text, raw_answer
        ))
        total += 1
        cats[category] = cats.get(category, 0) + 1
        types[qtype] = types.get(qtype, 0) + 1

    conn.commit()

    print(f"  ✅ {bank_name}: {total} 题")
    print(f"     分类: {json.dumps(cats, ensure_ascii=False)}")
    print(f"     题型: {json.dumps(types, ensure_ascii=False)}")

    # 验证前3题
    conn.row_factory = sqlite3.Row
    cur = conn.execute(
        "SELECT * FROM questions WHERE bank=? ORDER BY id DESC LIMIT 3", (bank_name,)
    )
    for r in cur.fetchall():
        print(f"     [{r['id']}] [{r['category']}] [{r['type']}] {r['content'][:50]}...")

    conn.close()
    return total


def convert():
    # 扫描目录下所有题库文件
    files = []
    for f in os.listdir(BASE_DIR):
        if f.endswith('.xls') or f.endswith('.xlsx'):
            files.append(f)

    if not files:
        print("[错误] 未找到任何 .xls / .xlsx 题库文件")
        return

    # 确保表存在
    conn = sqlite3.connect(DB_PATH)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank TEXT NOT NULL,
            category TEXT NOT NULL,
            type TEXT NOT NULL,
            content TEXT NOT NULL,
            options TEXT NOT NULL DEFAULT '[]',
            answer TEXT NOT NULL DEFAULT '[]',
            answer_text TEXT NOT NULL DEFAULT '',
            raw_answer TEXT NOT NULL DEFAULT ''
        )
    ''')
    conn.close()

    conn = sqlite3.connect(DB_PATH)
    existing_banks = set(row[0] for row in conn.execute("SELECT DISTINCT bank FROM questions").fetchall())
    conn.close()

    grand_total = 0
    for fname in files:
        filepath = os.path.join(BASE_DIR, fname)
        bank_name = os.path.splitext(fname)[0]
        bank_name = re.sub(r'^附件\d+[：:]', '', bank_name)
        if bank_name in existing_banks:
            print(f"\n⏭️  {fname} → 题库 [{bank_name}] 已存在，跳过")
            continue
        print(f"\n📄 {fname} → 题库名: {bank_name}")
        n = import_bank(bank_name, filepath)
        grand_total += n

    print(f"\n{'='*40}")
    print(f"🎉 全部导入完成！共 {grand_total} 道题")
    print(f"{'='*40}")


if __name__ == '__main__':
    convert()
